# Import Modules
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl import load_workbook, Workbook
from urllib.parse import urljoin, urlparse
import time
import re

def format_url(url):
    # Ensure the URL has a proper scheme
    if not url.startswith(('http://', 'https://')):
        url = "http://" + url
    return url

"""
Function to initialize WebDriver
- Setup Selenium WebDriver: Initialize the Selenium WebDriver.
"""
def init_driver():
    # Setup Chrome Options
    options = webdriver.ChromeOptions()
    #options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('log-level=3')

    # Initialize the Webdriver
    driver = webdriver.Chrome(options=options)
    return driver

"""
Function to load Excel File and Extract Data
"""
def load_excel_data(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook["Pre-Seed US"]

    companies_url_dict = {}

    for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row, min_col=3, max_col=7):
        company = row[0].value
        url = row[4].value
        if company  and url: 
            companies_url_dict[company] = url

    return companies_url_dict

def load_test_data(file_path):
    # Load the workbook and select the correct sheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active  # or specify the sheet name if you know it

    # Dictionary to store the company names and URLs
    companies_url_dict = {}

    # Iterate through the rows to extract company names and URLs
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        company = row[0].value  # Assuming the company name is in the first column (A)
        url = row[1].value  # Assuming the URL is in the second column (B)
        if company and url: 
            companies_url_dict[company] = url

    return companies_url_dict

"""
Function to extract all internal links from a single webpage
"""
def extract_internal_links(driver, base_url):
    internal_links = set()
    elements = driver.find_elements(By.TAG_NAME, 'a')

    for element in elements:
        href = element.get_attribute('href')
        if href:
            normalized_href = urljoin(base_url, href)
            if urlparse(normalized_href).netloc == urlparse(base_url).netloc:
                internal_links.add(normalized_href)
    
    return internal_links

"""
Function to recursively vist all the internal links on a site
"""
def traverse_site(driver, base_url):
    visited_links = set()
    links_to_visit = {base_url}
    search_results = {}

    while links_to_visit:
        current_link = links_to_visit.pop()
        if current_link in visited_links:
            continue # Skip already visited links

        print(f"Visiting: {current_link}")
        driver.get(current_link)
        visited_links.add(current_link)

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'title')))
            
            # Search for keywords in the current page
            page_search_results = search_keywords_in_elements(driver)
            if any(page_search_results.values()):
                search_results[current_link] = page_search_results

            # Extract Internal links and update the traversal list
            new_links = extract_internal_links(driver, base_url)
            links_to_visit.update(new_links - visited_links) # Only add unvisited links

        except TimeoutException:
            print(f"Timeout while accessing {current_link}")
        except Exception as e:
            print(f"Error while traversing {current_link}: {e}")
        
    return visited_links, search_results
        

def open_company_sites(driver, company_url_dict, output_file):
    search_data = {}
    for company, url in company_url_dict.items():
        search_data[company] = {"Site": url}

        try:
            # Format the URl correctly
            formmated_url = format_url(url)

            print(f"Accessing {company}: {formmated_url}")
            driver.get(formmated_url)
            
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'title')))
            _, search_results = traverse_site(driver, formmated_url)

            # Consolidate results into the search_data dictionary
            for page_url, keywords_found in search_results.items():
                for keyword, elements in keywords_found.items():
                    if elements:
                        if keyword not in search_data[company]:
                            search_data[company][keyword] = f"Yes; {page_url}; {elements[0]}"
                        else:
                            search_data[company][keyword] += f"; {page_url}; {elements[0]}"
            
        except TimeoutException:
            print(f"Timeout while accessing {company} at {url}")
        except Exception as e:
            print(f"Failed to access {company} at {url}: {e}")
    
    save_to_excel(search_data, output_file)

"""
Function to Search Keywords
- Logic to Scrape Websites: Load the page, extract links, and recursively crawl them.
- Search for Keywords: On each page, search for the given list of keywords.
"""
def search_keywords(page_content):
    keywords = [
        r'private\s*equity',
        r'capital\s*markets',
        r'leverage\s*finance', 
        r'investment\s*banking',
        r'investment\s*firm',
        r'b2b\s*saas',
        r'pre[-\s]*seed',
        r'southeast',
        r'latin',
        r'hispanic',
        r'florida'
    ]
    results = {}

    for keyword in keywords:
        if re.search(keyword, page_content, re.IGNORECASE):
            results[keyword] = "Yes"

    return results

def search_keywords_in_elements(driver):
    page_content = driver.page_source.lower()
    results = search_keywords(page_content)

    elements_results = {}
    for keyword, found in results.items():
        all_text_elements = driver.find_elements(By.XPATH, "//*[text()]")  # Find all elements with text

        matched_elements = []
        for element in all_text_elements:
            text = element.text.lower()
            if re.search(keyword, text):
                matched_elements.append(element.get_attribute('outerHTML'))
        
        if matched_elements:
            elements_results[keyword] = matched_elements

    return elements_results

"""
Function to save to an excel file
"""
def save_to_excel(data, output_file):
    # Open or create a workbook
    try:
        workbook = load_workbook(output_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        headers = ['Company', 'Site']
        sheet.append(headers)
    
    keyword_set = set()

    for company_data in data.values():
        for keyword in company_data.keys():
            if keyword != 'Site':
                keyword_set.add(keyword)
    
    # Extend headers with sorted keywords and append them if not already added
    current_headers = [cell.value for cell in sheet[1]]
    for keyword in sorted(keyword_set):
        if keyword not in current_headers:
            current_headers.append(keyword)
    
    sheet.delete_rows(1)  # Remove old header row
    sheet.append(current_headers)  # Append updated headers

    # Append each row of company data
    for company, results in data.items():
        row = [company, results.get('Site', '')]
        for keyword in sorted(keyword_set):
            row.append(results.get(keyword, ""))
        sheet.append(row)
    
    # Save the workbook after each row is added
    workbook.save(output_file)

"""
Main Scraper Function: The main function will load data, initialize the driver, iterate over each company URL, and use the keyword search function.
"""
def main_scraper(file_path, output_file):
    # Load the data from the excel file
    company_url_dict = load_test_data(file_path)

    driver = init_driver()

    # Open each company's website
    open_company_sites(driver, company_url_dict, output_file)

    driver.quit()

#### MAIN CODE ####
#file_path = r"C:\Users\dmaso\OneDrive\Documents\002 Projects\002 Freelance\1 - Javier Gutierrez\Scraper00\data\Aug 28 - VCs.xlsx"
file_path = r"C:\Users\dmaso\OneDrive\Documents\002 Projects\002 Freelance\1 - Javier Gutierrez\Scraper00\data\Test-Data.xlsx"
output_file = r"C:\Users\dmaso\OneDrive\Documents\002 Projects\002 Freelance\1 - Javier Gutierrez\Scraper00\results.xlsx"
"""
# Test the load_excel_file function
companies_url_dict = load_excel_data(file_path)

print(f"Companies: {companies_url_dict}")
"""

main_scraper(file_path, output_file)